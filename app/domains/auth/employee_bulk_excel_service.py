from datetime import date, datetime
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError

from app.models.wms import UserRole
from app.domains.auth.schemas.auth import (
    EmployeeBulkCreateResultRow,
    EmployeeBulkCreateRow,
)


EMPLOYEE_BULK_INPUT_HEADERS = (
    "이름",
    "입사일",
    "역할",
    "이메일",
)

EMPLOYEE_BULK_RESULT_HEADERS = (
    "이름",
    "입사일",
    "역할",
    "이메일",
    "사번",
    "최초 비밀번호",
)

MAX_EMPLOYEE_BULK_CREATE_ROWS = 500

ROLE_VALUE_MAP = {
    "ADMIN": UserRole.ADMIN,
    "관리자": UserRole.ADMIN,
    "WORKER": UserRole.WORKER,
    "작업자": UserRole.WORKER,
}


class EmployeeBulkExcelValidationError(ValueError):
    """
    엑셀 형식 또는 개별 행 검증 오류.

    errors에는 관리자 화면에서 그대로 보여줄 수 있는
    행 단위 오류 메시지를 저장한다.
    """

    def __init__(
        self,
        errors: list[str],
    ) -> None:
        self.errors = errors

        super().__init__(
            "직원 일괄 생성 엑셀 검증에 실패했습니다."
        )


def _normalize_cell_text(
    value: object | None,
) -> str | None:
    if value is None:
        return None

    normalized = str(value).strip()

    return normalized or None


def _parse_hire_date(
    value: object | None,
) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None

    return None


def _parse_role(
    value: object | None,
) -> UserRole | None:
    role_text = _normalize_cell_text(value)

    if role_text is None:
        return None

    return ROLE_VALUE_MAP.get(
        role_text.upper(),
    )


def _format_validation_error(
    *,
    source_row: int,
    exc: ValidationError,
) -> str:
    messages = []

    for error in exc.errors():
        field_name = (
            str(error["loc"][-1])
            if error.get("loc")
            else "알 수 없는 필드"
        )
        messages.append(
            f"{field_name}: {error['msg']}"
        )

    return (
        f"{source_row}행 - "
        + ", ".join(messages)
    )


def parse_employee_bulk_create_xlsx(
    file_content: bytes,
) -> list[EmployeeBulkCreateRow]:
    """
    업로드된 직원 일괄 생성 엑셀을 검증된 직원 행 목록으로 변환한다.

    입력 헤더는 반드시:
    이름 | 입사일 | 역할 | 이메일
    순서여야 한다.
    """
    try:
        workbook = load_workbook(
            filename=BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except (
        BadZipFile,
        InvalidFileException,
        KeyError,
        OSError,
    ) as exc:
        raise EmployeeBulkExcelValidationError(
            ["업로드 파일은 올바른 .xlsx 형식이어야 합니다."]
        ) from exc

    try:
        worksheet = workbook.active

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if header_row is None:
            raise EmployeeBulkExcelValidationError(
                ["엑셀 헤더 행이 없습니다."]
            )

        actual_headers = tuple(
            _normalize_cell_text(value)
            for value in header_row[:len(
                EMPLOYEE_BULK_INPUT_HEADERS
            )]
        )

        if actual_headers != EMPLOYEE_BULK_INPUT_HEADERS:
            expected = " | ".join(
                EMPLOYEE_BULK_INPUT_HEADERS
            )
            actual = " | ".join(
                value or "(비어 있음)"
                for value in actual_headers
            )

            raise EmployeeBulkExcelValidationError(
                [
                    "엑셀 헤더가 올바르지 않습니다. "
                    f"기대값: {expected}, 실제값: {actual}"
                ]
            )

        parsed_rows: list[
            EmployeeBulkCreateRow
        ] = []
        errors: list[str] = []

        for source_row, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            input_values = values[:len(
                EMPLOYEE_BULK_INPUT_HEADERS
            )]

            if all(
                value is None
                or _normalize_cell_text(value) is None
                for value in input_values
            ):
                continue

            name = _normalize_cell_text(
                input_values[0],
            )
            hire_date = _parse_hire_date(
                input_values[1],
            )
            role = _parse_role(
                input_values[2],
            )
            email = _normalize_cell_text(
                input_values[3],
            )

            if input_values[1] is not None and hire_date is None:
                errors.append(
                    f"{source_row}행 - 입사일은 "
                    "YYYY-MM-DD 형식이어야 합니다."
                )
                continue

            if input_values[2] is not None and role is None:
                errors.append(
                    f"{source_row}행 - 역할은 ADMIN, WORKER, "
                    "관리자, 작업자 중 하나여야 합니다."
                )
                continue

            try:
                parsed_rows.append(
                    EmployeeBulkCreateRow(
                        source_row=source_row,
                        name=name,
                        hire_date=hire_date,
                        role=role,
                        email=email,
                    )
                )
            except ValidationError as exc:
                errors.append(
                    _format_validation_error(
                        source_row=source_row,
                        exc=exc,
                    )
                )

        if not parsed_rows and not errors:
            errors.append(
                "생성할 직원 행이 없습니다."
            )

        if (
            len(parsed_rows)
            > MAX_EMPLOYEE_BULK_CREATE_ROWS
        ):
            errors.append(
                "한 번에 생성할 수 있는 직원 수는 "
                f"{MAX_EMPLOYEE_BULK_CREATE_ROWS}명입니다."
            )

        if errors:
            raise EmployeeBulkExcelValidationError(
                errors
            )

        return parsed_rows

    finally:
        workbook.close()


def _apply_header_style(
    worksheet,
) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font


def build_employee_bulk_template_xlsx() -> bytes:
    """
    MASTER가 다운로드하여 작성할 직원 일괄 생성 양식 엑셀을 생성한다.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "직원 일괄 생성"

    worksheet.append(
        EMPLOYEE_BULK_INPUT_HEADERS
    )
    _apply_header_style(worksheet)

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 32

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def build_employee_bulk_result_xlsx(
    results: list[EmployeeBulkCreateResultRow],
) -> bytes:
    """
    계정 생성 결과와 최초 비밀번호를 포함한 다운로드용 엑셀을 생성한다.

    최초 비밀번호는 이 결과 파일에서만 제공하며 DB에는 저장하지 않는다.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "직원 계정 발급 결과"

    worksheet.append(
        EMPLOYEE_BULK_RESULT_HEADERS
    )
    _apply_header_style(worksheet)

    for result in results:
        worksheet.append(
            [
                result.name,
                result.hire_date,
                result.role.value,
                result.email,
                result.employee_id,
                result.temporary_password,
            ]
        )

    worksheet.freeze_panes = "A2"

    for column, width in {
        "A": 18,
        "B": 16,
        "C": 14,
        "D": 32,
        "E": 18,
        "F": 24,
    }.items():
        worksheet.column_dimensions[
            column
        ].width = width

    for cell in worksheet["B"][1:]:
        cell.number_format = "yyyy-mm-dd"

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()