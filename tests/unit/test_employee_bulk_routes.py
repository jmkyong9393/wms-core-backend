from datetime import date
from io import BytesIO
from types import SimpleNamespace
from uuid import UUID

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.dependencies.auth import require_master
from app.api.routes import admin_users
from app.core.database import get_session
from app.models.wms import UserRole
from app.schemas.auth import (
    EmployeeBulkCreateResultRow,
    EmployeeBulkCreateRow,
)
from app.services.employee_bulk_excel_service import (
    EmployeeBulkExcelValidationError,
)


@pytest.fixture
def client():
    app = FastAPI()

    app.include_router(
        admin_users.router,
        prefix="/api/v1/users/admin",
    )

    current_master = SimpleNamespace(
        id=UUID(
            "00000000-0000-4000-8000-000000000001"
        ),
        tenant_id=UUID(
            "00000000-0000-4000-8000-000000000100"
        ),
        employee_id="NZ0000000",
    )

    app.dependency_overrides[
        require_master
    ] = lambda: current_master
    app.dependency_overrides[
        get_session
    ] = lambda: SimpleNamespace()

    with TestClient(app) as test_client:
        yield test_client

    app.dependency_overrides.clear()


def test_downloads_bulk_employee_template(client):
    response = client.get(
        "/api/v1/users/admin/bulk-template"
    )

    assert response.status_code == 200
    assert response.headers[
        "content-type"
    ].startswith(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    assert "attachment" in response.headers[
        "content-disposition"
    ]
    assert response.headers["cache-control"] == "no-store"

    workbook = load_workbook(
        BytesIO(response.content),
        data_only=True,
    )
    worksheet = workbook.active

    assert [
        cell.value
        for cell in worksheet[1]
    ] == [
        "이름",
        "입사일",
        "역할",
        "이메일",
    ]

    workbook.close()


def test_bulk_create_returns_result_xlsx(
    client,
    monkeypatch,
):
    input_row = EmployeeBulkCreateRow(
        source_row=2,
        name="김가나",
        hire_date=date(2026, 8, 1),
        role=UserRole.WORKER,
        email="gana@example.com",
    )

    result_row = EmployeeBulkCreateResultRow(
        source_row=2,
        name="김가나",
        hire_date=date(2026, 8, 1),
        role=UserRole.WORKER,
        email="gana@example.com",
        employee_id="NZ2608001",
        temporary_password="TempPassword123!",
    )

    monkeypatch.setattr(
        admin_users,
        "parse_employee_bulk_create_xlsx",
        lambda _file_content: [input_row],
    )
    monkeypatch.setattr(
        admin_users,
        "create_employees_bulk",
        lambda **_kwargs: [result_row],
    )

    response = client.post(
        "/api/v1/users/admin/bulk-create",
        files={
            "file": (
                "employees.xlsx",
                b"test-xlsx-content",
                (
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
            )
        },
    )

    assert response.status_code == 201
    assert response.headers[
        "content-type"
    ].startswith(
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
    assert "attachment" in response.headers[
        "content-disposition"
    ]
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["pragma"] == "no-cache"

    workbook = load_workbook(
        BytesIO(response.content),
        data_only=True,
    )
    worksheet = workbook.active

    assert worksheet["A2"].value == "김가나"
    assert worksheet["E2"].value == "NZ2608001"
    assert worksheet["F2"].value == "TempPassword123!"

    workbook.close()


def test_bulk_create_returns_row_errors_for_invalid_xlsx(
    client,
    monkeypatch,
):
    def raise_validation_error(_file_content):
        raise EmployeeBulkExcelValidationError(
            [
                "2행 - 역할은 ADMIN, WORKER, 관리자, "
                "작업자 중 하나여야 합니다."
            ]
        )

    monkeypatch.setattr(
        admin_users,
        "parse_employee_bulk_create_xlsx",
        raise_validation_error,
    )

    response = client.post(
        "/api/v1/users/admin/bulk-create",
        files={
            "file": (
                "employees.xlsx",
                b"invalid-xlsx-content",
                (
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"]["errors"] == [
        "2행 - 역할은 ADMIN, WORKER, 관리자, "
        "작업자 중 하나여야 합니다."
    ]


def test_bulk_create_rejects_non_xlsx_file(client):
    response = client.post(
        "/api/v1/users/admin/bulk-create",
        files={
            "file": (
                "employees.csv",
                b"name,hire_date,role,email",
                "text/csv",
            )
        },
    )

    assert response.status_code == 422
    assert "xlsx" in response.json()["detail"][
        "message"
    ]