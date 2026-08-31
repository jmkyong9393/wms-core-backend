from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.models.wms import UserRole
from app.domains.auth.schemas.auth import EmployeeBulkCreateResultRow
from app.domains.auth.employee_bulk_excel_service import (
    EMPLOYEE_BULK_INPUT_HEADERS,
    EmployeeBulkExcelValidationError,
    build_employee_bulk_result_xlsx,
    parse_employee_bulk_create_xlsx,
)


def build_input_xlsx(
    rows: list[list[object]],
    headers: tuple[str, ...] = EMPLOYEE_BULK_INPUT_HEADERS,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(headers)

    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)
    workbook.close()

    return output.getvalue()


def test_parses_valid_employee_bulk_xlsx():
    content = build_input_xlsx(
        [
            [
                "김가나",
                date(2026, 8, 1),
                "WORKER",
                "gana@example.com",
            ],
            [
                "나다라",
                "2026-08-02",
                "관리자",
                None,
            ],
        ]
    )

    rows = parse_employee_bulk_create_xlsx(content)

    assert len(rows) == 2

    assert rows[0].source_row == 2
    assert rows[0].name == "김가나"
    assert rows[0].hire_date == date(2026, 8, 1)
    assert rows[0].role == UserRole.WORKER
    assert str(rows[0].email) == "gana@example.com"

    assert rows[1].source_row == 3
    assert rows[1].hire_date == date(2026, 8, 2)
    assert rows[1].role == UserRole.ADMIN
    assert rows[1].email is None


def test_rejects_invalid_bulk_excel_headers():
    content = build_input_xlsx(
        rows=[],
        headers=(
            "이름",
            "생성일",
            "직책",
            "이메일",
        ),
    )

    with pytest.raises(
        EmployeeBulkExcelValidationError,
    ) as exc_info:
        parse_employee_bulk_create_xlsx(content)

    assert any(
        "헤더" in error
        for error in exc_info.value.errors
    )


def test_collects_invalid_role_and_date_errors():
    content = build_input_xlsx(
        [
            [
                "김가나",
                "2026/08/01",
                "WORKER",
                "gana@example.com",
            ],
            [
                "나다라",
                "2026-08-02",
                "MANAGER",
                "dara@example.com",
            ],
        ]
    )

    with pytest.raises(
        EmployeeBulkExcelValidationError,
    ) as exc_info:
        parse_employee_bulk_create_xlsx(content)

    errors = exc_info.value.errors

    assert any("2행" in error for error in errors)
    assert any("입사일" in error for error in errors)
    assert any("3행" in error for error in errors)
    assert any("역할" in error for error in errors)


def test_builds_result_xlsx_with_issued_credentials():
    content = build_employee_bulk_result_xlsx(
        [
            EmployeeBulkCreateResultRow(
                source_row=2,
                name="김가나",
                hire_date=date(2026, 8, 1),
                role=UserRole.WORKER,
                email="gana@example.com",
                employee_id="NZ2608001",
                temporary_password="TempPassword123!",
            )
        ]
    )

    workbook = load_workbook(
        BytesIO(content),
        data_only=True,
    )
    worksheet = workbook.active

    assert [
        cell.value
        for cell in worksheet[1]
    ] == [
        "이름",
        "입사일",
        "역할",
        "이메일",
        "사번",
        "최초 비밀번호",
    ]

    result_row = [
        cell.value
        for cell in worksheet[2]
    ]

    assert result_row[0] == "김가나"
    assert result_row[2] == "WORKER"
    assert result_row[4] == "NZ2608001"
    assert result_row[5] == "TempPassword123!"

    workbook.close()
    